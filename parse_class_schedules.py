"""
One-time parser that extracts the academic class schedules from the files in
"מערכות שעות/" and writes them to class_schedules.json.

Run:
    python parse_class_schedules.py

The output JSON has the structure:
{
  "<institution>": {
    "<track>": {
      "<year>": [
        {"day": "ראשון", "start": "09:00", "end": "11:00",
         "course": "...", "type": "handasa" | "handasai"}
      ]
    }
  }
}

Sources handled automatically:
  * Sami Shamoon (SCE) Excel workbook (one sheet per class).
  * Braude electronics PDFs.

Ariel JPEGs are not OCR'd here. Their data lives in
ariel_class_schedules.json (hand-curated) and is merged in.
"""

import json
import os
import re
import shutil
import tempfile
from collections import defaultdict

import openpyxl
import pdfplumber

WEEKDAYS = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי"]
HEB_LETTER_TO_DAY = {"א": "ראשון", "ב": "שני", "ג": "שלישי",
                     "ד": "רביעי", "ה": "חמישי", "ו": "שישי"}

ROOT = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(ROOT, "מערכות שעות")
OUTPUT_FILE = os.path.join(ROOT, "class_schedules.json")
ARIEL_FILE = os.path.join(ROOT, "ariel_class_schedules.json")


# ────────────────────────────────────────────────────────────────
# Sami Shamoon (SCE)
# ────────────────────────────────────────────────────────────────

SCE_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})")
YELLOW_FILLS = {"FFFFFF00", "FFFFFFCC", "FFFFFFA0"}  # variations of yellow


def _normalise_rgb(rgb):
    if not rgb:
        return None
    if isinstance(rgb, str):
        return rgb.upper()
    return None


def _is_yellow(cell):
    fg = cell.fill.fgColor if cell.fill else None
    rgb = _normalise_rgb(fg.rgb if fg and fg.type == "rgb" else None)
    if not rgb:
        return False
    # Yellow has high R+G, low B.  Heuristic catches the typical FFFFFF00 etc.
    if rgb in YELLOW_FILLS:
        return True
    try:
        r = int(rgb[-6:-4], 16)
        g = int(rgb[-4:-2], 16)
        b = int(rgb[-2:], 16)
        return r > 200 and g > 200 and b < 100
    except ValueError:
        return False


def _resolve_merged(ws, row, col):
    """Return (top_row, bottom_row, anchor_cell) for the merged range that
    contains (row, col), or (row, row, cell) if not merged."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.max_row, ws.cell(rng.min_row, rng.min_col)
    return row, row, ws.cell(row, col)


def _row_to_hour(ws, row, header_row=2, time_col=3):
    """Translate a row number into HH:MM based on the שעות column."""
    val = ws.cell(row, time_col).value
    if hasattr(val, "hour"):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, str):
        m = re.match(r"(\d{1,2}):(\d{2})", val)
        if m:
            return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    return None


def parse_sce(xlsx_path):
    """Return {track: {year: [events]}} for the SCE workbook."""
    # The live file is often open in Excel; copy to a temp location.
    with tempfile.TemporaryDirectory() as tmp:
        local = os.path.join(tmp, "sce.xlsx")
        shutil.copy(xlsx_path, local)
        wb = openpyxl.load_workbook(local, data_only=True)

        result = defaultdict(lambda: defaultdict(list))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            track, year = _split_sheet_name(sheet_name)
            if not track:
                continue

            # Header row (row 2): col 3=שעות, col 4..=days א..ו
            day_cols = {}
            for c in range(4, ws.max_column + 1):
                v = ws.cell(2, c).value
                if isinstance(v, str):
                    for letter, name in HEB_LETTER_TO_DAY.items():
                        if v.strip().endswith(letter):
                            day_cols[c] = name
                            break

            seen_blocks = set()  # (row_top, col) for merged anchors
            for row in range(3, ws.max_row + 1):
                for col, day_name in day_cols.items():
                    top, bot, anchor = _resolve_merged(ws, row, col)
                    key = (top, col)
                    if key in seen_blocks:
                        continue
                    seen_blocks.add(key)

                    raw = anchor.value
                    if raw is None or (isinstance(raw, str) and not raw.strip()):
                        continue

                    text = str(raw).replace("\r", "")
                    yellow = _is_yellow(anchor)

                    # Try to find an explicit time range in the text.
                    m = SCE_TIME_RE.search(text)
                    if m:
                        sh, sm, eh, em = (int(g) for g in m.groups())
                        start = f"{sh:02d}:{sm:02d}"
                        end = f"{eh:02d}:{em:02d}"
                        # Remove the time substring from the course name.
                        course = SCE_TIME_RE.sub("", text).strip(" \n\t-:")
                    else:
                        start = _row_to_hour(ws, top) or "00:00"
                        end = _row_to_hour(ws, bot + 1) or _row_to_hour(ws, bot) or start
                        course = text.strip()

                    if yellow:
                        course = course if course and not SCE_TIME_RE.fullmatch(course) else "Handasai Course"
                        ev_type = "handasai"
                    else:
                        ev_type = "handasa"

                    course = re.sub(r"\s+", " ", course).strip()
                    if not course:
                        course = "Handasai Course" if yellow else "קורס"

                    result[track][year].append({
                        "day": day_name,
                        "start": start,
                        "end": end,
                        "course": course,
                        "type": ev_type,
                    })

        return {k: dict(v) for k, v in result.items()}


def _split_sheet_name(name):
    """'תוכנה שנה א' → ('תוכנה', 'א')."""
    m = re.match(r"^(.*?)\s*שנה\s*([אבגדה])\s*$", name.strip())
    if not m:
        return None, None
    return m.group(1).strip(), m.group(2)


# ────────────────────────────────────────────────────────────────
# Braude (PDF)
# ────────────────────────────────────────────────────────────────

HEBREW_CHARS_RE = re.compile(r"[\u0590-\u05FF]")
TIME_RANGE_RE = re.compile(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")


def _reverse_hebrew(text):
    """pdfplumber returns visual order for RTL text; flip Hebrew words."""
    if HEBREW_CHARS_RE.search(text):
        return text[::-1]
    return text


def parse_braude(pdf_path, track, year):
    """Return list of events for a single Braude PDF.

    Layout (per visual inspection): rows are grouped vertically into
    period bands.  Within each band you have, from top to bottom:
        ~21 above time : course-name row (Hebrew words across all 5 days)
        ~ 9 above time : teacher + room (optional)
                  time : "HH:MM-HH:MM" — one per day column
    Continuation periods have the word "המשך" on the course-name row and
    must be ignored per spec.
    """
    events = []
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        words = page.extract_words(use_text_flow=False)
        if not words:
            return events

        # ── Day column centres from the date words in the header row ──
        date_words = [w for w in words if re.match(r"^\d{2}/\d{2}/\d{2,4}$", w["text"])]
        date_words.sort(key=lambda w: -w["x0"])  # rightmost = יום א
        if len(date_words) < 5:
            return events
        day_centres = [(w["x0"] + w["x1"]) / 2 for w in date_words[:5]]
        day_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי"]

        def column_for(x):
            best = min(range(5), key=lambda i: abs(day_centres[i] - x))
            return best if abs(day_centres[best] - x) < 45 else None

        # ── Cluster words into rows by top (3-pt tolerance) ──
        rows = defaultdict(list)
        for w in words:
            rows[round(w["top"] / 3) * 3].append(w)
        row_tops = sorted(rows.keys())

        # ── For each row that contains time-range words, gather data ──
        for top in row_tops:
            row_words = rows[top]
            time_words = [w for w in row_words if TIME_RANGE_RE.match(w["text"])]
            if len(time_words) < 2:  # need at least a couple to count as a slot row
                continue

            # Find the course-name row above (~21 pts up, ±6).
            name_row_top = None
            for cand in row_tops:
                if 14 <= top - cand <= 28:
                    name_row_top = cand
                    break
            if name_row_top is None:
                continue
            name_words = rows[name_row_top]

            # If the row above is the "המשך" continuation row, skip it.
            if any(w["text"] == "ךשמה" or "המשך" in w["text"] for w in name_words):
                continue

            # Per-day extraction
            for tw in time_words:
                m = TIME_RANGE_RE.match(tw["text"])
                sh, sm, eh, em = (int(g) for g in m.groups())
                start = f"{sh:02d}:{sm:02d}"
                end = f"{eh:02d}:{em:02d}"
                col = column_for((tw["x0"] + tw["x1"]) / 2)
                if col is None:
                    continue
                day = day_names[col]
                cx_min = day_centres[col] - 40
                cx_max = day_centres[col] + 40

                in_col = [
                    w for w in name_words
                    if cx_min <= (w["x0"] + w["x1"]) / 2 <= cx_max
                    and HEBREW_CHARS_RE.search(w["text"])
                    and w["text"] != "ךשמה"
                ]
                in_col.sort(key=lambda w: -w["x0"])  # RTL
                course = " ".join(_reverse_hebrew(w["text"]) for w in in_col).strip()
                if not course:
                    continue
                events.append({
                    "day": day,
                    "start": start,
                    "end": end,
                    "course": course,
                    "type": "handasa",
                })

    # Merge contiguous slots of the same course on the same day.
    events.sort(key=lambda e: (e["day"], e["course"], e["start"]))
    merged = []
    for ev in events:
        if (merged and merged[-1]["day"] == ev["day"]
                and merged[-1]["course"] == ev["course"]
                and merged[-1]["end"] >= ev["start"]):
            merged[-1]["end"] = max(merged[-1]["end"], ev["end"])
        else:
            merged.append(dict(ev))
    return merged


# ────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────

def main():
    out = {
        "סמי שמעון": {},
        "בראודה": {},
        "אוניברסיטת אריאל": {},
    }

    # ── SCE ────────────────────────────────────────────────────────
    sce_path = os.path.join(SOURCE_DIR, "מערכות שעות סמי שמעון.xlsx")
    if os.path.exists(sce_path):
        try:
            out["סמי שמעון"] = parse_sce(sce_path)
        except PermissionError:
            print(f"WARNING: {sce_path} is locked (close it in Excel). Keeping previous SCE data if present.")
            if os.path.exists(OUTPUT_FILE):
                with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
                    prev = json.load(f)
                out["סמי שמעון"] = prev.get("סמי שמעון", {})

    # ── Braude ─────────────────────────────────────────────────────
    braude_dir = os.path.join(SOURCE_DIR, "בראודה")
    if os.path.isdir(braude_dir):
        track = "חשמל"  # אלקטרוניקה sits under the חשמל track at Braude
        out["בראודה"].setdefault(track, {})
        for fname in os.listdir(braude_dir):
            if not fname.lower().endswith(".pdf"):
                continue
            m = re.search(r"שנה\s*([אבגדה])", fname)
            year = m.group(1) if m else "א"
            evs = parse_braude(os.path.join(braude_dir, fname), track, year)
            out["בראודה"][track][year] = evs

    # ── Ariel (manual JSON) ────────────────────────────────────────
    if os.path.exists(ARIEL_FILE):
        with open(ARIEL_FILE, "r", encoding="utf-8") as f:
            ariel_data = json.load(f)
        # Strip meta keys (start with "_")
        out["אוניברסיטת אריאל"] = {
            k: v for k, v in ariel_data.items()
            if not k.startswith("_") and isinstance(v, dict)
        }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    total = 0
    for inst in out.values():
        for track in inst.values():
            if not isinstance(track, dict):
                continue
            for evs in track.values():
                if isinstance(evs, list):
                    total += len(evs)
    print(f"Wrote {OUTPUT_FILE} ({total} events)")


if __name__ == "__main__":
    main()
