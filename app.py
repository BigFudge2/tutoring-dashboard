"""Flask dashboard למעקב תגבורים."""
from __future__ import annotations

import json
import io
from datetime import datetime, timedelta
from typing import Any
import os

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_caching import Cache

import config
from data_loader import (
    COL_DATE,
    COL_END,
    COL_NOTES,
    COL_RATING,
    COL_START,
    COL_STUDENTS,
    COL_TIMESTAMP,
    COL_TOPIC,
    COL_TUTOR,
    append_row,
    get_all_students,
    get_unique_students,
    get_unique_tutors,
    load_data,
    load_probation_students,
    append_probation_student,
    remove_probation_student,
    load_tutors_registry,
    append_tutor,
    update_tutor,
    remove_tutor,
    get_tutor_subjects,
    load_weekly_schedule,
    append_schedule_slot,
    update_schedule_slot,
    remove_schedule_slot,
    load_onetime_lessons,
    append_onetime_lesson,
    update_onetime_lesson,
    remove_onetime_lesson,
    remove_schedule_by_tutor,
    load_registrations,
)
from topic_matcher import build_topic_map, group_similar_topics


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
cache = Cache(app, config={"CACHE_TYPE": "SimpleCache", "CACHE_DEFAULT_TIMEOUT": config.CACHE_TTL_SECONDS})

# ---------- נתיבים ציבוריים (ללא הגנה) ----------
PUBLIC_PREFIXES = ("/form", "/api/submit", "/login", "/static")


@app.before_request
def require_login():
    """דורש התחברות לכל הנתיבים חוץ מטופס דיווח ו-login."""
    if any(request.path.startswith(p) for p in PUBLIC_PREFIXES):
        return None
    if session.get("admin"):
        return None
    return redirect(url_for("login_page"))


@app.route("/login", methods=["GET", "POST"])
def login_page():
    error = None
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect("/")
        error = "סיסמה שגויה"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect(url_for("login_page"))


# ---------- Class schedules (academic overlay) ----------

CLASS_SCHEDULES_FILE = os.path.join(os.path.dirname(__file__), "class_schedules.json")
_class_schedules_cache: dict[str, Any] | None = None
_class_schedules_mtime: float = 0.0


def _load_class_schedules() -> dict[str, Any]:
    """Load and cache class_schedules.json, reloading if the file changes."""
    global _class_schedules_cache, _class_schedules_mtime
    if not os.path.exists(CLASS_SCHEDULES_FILE):
        return {}
    mtime = os.path.getmtime(CLASS_SCHEDULES_FILE)
    if _class_schedules_cache is None or mtime != _class_schedules_mtime:
        with open(CLASS_SCHEDULES_FILE, "r", encoding="utf-8") as f:
            _class_schedules_cache = json.load(f)
        _class_schedules_mtime = mtime
    return _class_schedules_cache or {}


def _get_academic_events(institution: str | None, track: str | None,
                         year: str | None = None) -> list[dict[str, Any]]:
    """Return academic class events for the given institution+track filter.

    Returns an empty list unless BOTH filters are provided. If `year` is
    also supplied, only that year's events are returned.
    """
    if not institution or not track:
        return []
    data = _load_class_schedules()
    track_data = data.get(institution, {}).get(track, {})
    if not isinstance(track_data, dict):
        return []
    events: list[dict[str, Any]] = []
    for yr, evs in track_data.items():
        if not isinstance(evs, list):
            continue
        if year and yr != year:
            continue
        for ev in evs:
            events.append({
                "day": ev.get("day"),
                "start": ev.get("start"),
                "end": ev.get("end"),
                "course": ev.get("course", ""),
                "type": ev.get("type", "handasa"),
                "year": yr,
                "kind": "academic",
            })
    return events


def _available_years(institution: str | None, track: str | None) -> list[str]:
    if not institution or not track:
        return []
    data = _load_class_schedules()
    track_data = data.get(institution, {}).get(track, {})
    if not isinstance(track_data, dict):
        return []
    return [y for y, evs in track_data.items() if isinstance(evs, list) and evs]


# ---------- Helpers ----------

def _get_df() -> pd.DataFrame:
    cached = cache.get("df")
    if cached is not None:
        return cached
    df = load_data()
    if not df.empty and COL_TOPIC in df.columns:
        topic_map = build_topic_map(
            df[COL_TOPIC].dropna().tolist(),
            threshold=config.TOPIC_SIMILARITY_THRESHOLD,
        )
        df["_topic_normalized"] = df[COL_TOPIC].map(lambda t: topic_map.get(t, t))
    elif not df.empty:
        df["_topic_normalized"] = ""
    cache.set("df", df)
    return df


def _row_to_dict(row: pd.Series) -> dict[str, Any]:
    """ממיר שורה ל-dict ידידותי ל-Jinja."""
    d: dict[str, Any] = {}
    for k, v in row.items():
        if str(k).startswith("_"):
            continue
        if pd.isna(v):
            d[str(k)] = ""
        elif isinstance(v, pd.Timestamp):
            d[str(k)] = v.strftime("%d/%m/%Y")
        elif isinstance(v, datetime):
            d[str(k)] = v.strftime("%d/%m/%Y")
        else:
            d[str(k)] = v
    return d


def _df_to_records(df: pd.DataFrame) -> list[dict[str, Any]]:
    return [_row_to_dict(row) for _, row in df.iterrows()]


# ---------- Cached data helpers ----------

def _get_registry() -> list[dict]:
    cached = cache.get("registry")
    if cached is not None:
        return cached
    data = load_tutors_registry()
    cache.set("registry", data)
    return data


def _get_weekly() -> list[dict]:
    cached = cache.get("weekly")
    if cached is not None:
        return cached
    data = load_weekly_schedule()
    cache.set("weekly", data)
    return data


def _get_onetime() -> list[dict]:
    cached = cache.get("onetime")
    if cached is not None:
        return cached
    data = load_onetime_lessons()
    cache.set("onetime", data)
    return data


# ---------- Routes ----------

@app.route("/")
def index():
    return redirect("/schedule")


@app.route("/tutors")
def tutors_page():
    df = _get_df()
    all_tutors = get_unique_tutors(df)
    selected = request.args.get("name") or (all_tutors[0] if all_tutors else "")
    tutor_df = df[df[COL_TUTOR] == selected].copy() if COL_TUTOR in df.columns and selected else pd.DataFrame()

    stats = {
        "lessons": len(tutor_df),
        "hours": round(float(tutor_df["משך (שעות)"].sum()), 1)
        if "משך (שעות)" in tutor_df.columns and not tutor_df.empty
        else 0.0,
        "avg_duration": round(float(tutor_df["משך (שעות)"].mean()), 2)
        if "משך (שעות)" in tutor_df.columns and not tutor_df.empty
        else 0.0,
        "avg_rating": round(float(tutor_df[COL_RATING].dropna().mean()), 2)
        if COL_RATING in tutor_df.columns and tutor_df[COL_RATING].notna().any()
        else None,
        "students_count": len({s for lst in tutor_df.get("_students_list", []) for s in lst}) if not tutor_df.empty else 0,
    }

    weekly_labels: list[str] = []
    weekly_hours: list[float] = []
    if not tutor_df.empty and COL_DATE in tutor_df.columns and "משך (שעות)" in tutor_df.columns:
        weekly = (
            tutor_df.dropna(subset=[COL_DATE])
            .assign(week=lambda d: d[COL_DATE].dt.to_period("W").astype(str))
            .groupby("week")["משך (שעות)"]
            .sum()
            .sort_index()
        )
        weekly_labels = weekly.index.tolist()
        weekly_hours = [round(float(v), 2) for v in weekly.values]

    if not tutor_df.empty and "_topic_normalized" in tutor_df.columns:
        topic_counts = tutor_df["_topic_normalized"].value_counts()
        topic_labels = topic_counts.index.tolist()
        topic_values = [int(v) for v in topic_counts.values]
    else:
        topic_labels, topic_values = [], []

    # Tutor registry record (subjects + probation students assignment).
    tutors_registry = cache.get("tutors_registry")
    if tutors_registry is None:
        tutors_registry = load_tutors_registry()
        cache.set("tutors_registry", tutors_registry)
    tutor_record = next((t for t in tutors_registry if t["name"] == selected), None) or {}

    return render_template(
        "tutors.html",
        tutors=all_tutors,
        selected=selected,
        stats=stats,
        weekly_labels=json.dumps(weekly_labels, ensure_ascii=False),
        weekly_hours=json.dumps(weekly_hours),
        topic_labels=json.dumps(topic_labels, ensure_ascii=False),
        topic_values=json.dumps(topic_values),
        lessons=_df_to_records(tutor_df),
        tutor_record=tutor_record,
    )


@app.route("/topics")
def topics_page():
    df = _get_df()
    threshold = int(request.args.get("threshold", config.TOPIC_SIMILARITY_THRESHOLD))
    topics_list = df[COL_TOPIC].dropna().tolist() if COL_TOPIC in df.columns else []
    groups = group_similar_topics(topics_list, threshold=threshold)

    topic_map = {v: rep for rep, variants in groups.items() for v in variants}
    rows = []
    for rep, variants in groups.items():
        if COL_TOPIC in df.columns:
            mask = df[COL_TOPIC].map(lambda t: topic_map.get(t, t)) == rep
            count = int(mask.sum())
            tutors_list = df.loc[mask, COL_TUTOR].dropna().unique().tolist() if COL_TUTOR in df.columns else []
        else:
            count, tutors_list = 0, []
        rows.append({
            "topic": rep,
            "variants": [v for v in variants if v != rep],
            "lessons": count,
            "tutors": tutors_list,
        })
    rows.sort(key=lambda r: r["lessons"], reverse=True)

    chart_labels = [r["topic"] for r in rows[:15]]
    chart_values = [r["lessons"] for r in rows[:15]]

    return render_template(
        "topics.html",
        threshold=threshold,
        groups=rows,
        total_groups=len(rows),
        chart_labels=json.dumps(chart_labels, ensure_ascii=False),
        chart_values=json.dumps(chart_values),
    )


@app.route("/students")
def students_page():
    df = _get_df()
    students = get_unique_students(df)
    selected = request.args.get("name")

    # אם לא נבחר סטודנט – מציג סקירת כלל הסטודנטים
    if not selected:
        all_students_stats = []
        for name in students:
            if "_students_list" in df.columns:
                mask = df["_students_list"].apply(lambda lst, n=name: n in lst)
                student_df = df[mask]
            else:
                student_df = pd.DataFrame()
            all_students_stats.append({
                "name": name,
                "lessons": len(student_df),
                "hours": round(float(student_df["משך (שעות)"].sum()), 1)
                if "משך (שעות)" in student_df.columns and not student_df.empty
                else 0.0,
                "tutors_count": student_df[COL_TUTOR].nunique()
                if COL_TUTOR in student_df.columns and not student_df.empty
                else 0,
            })
        all_students_stats.sort(key=lambda s: s["lessons"])
        return render_template(
            "students.html",
            students=students,
            selected=None,
            all_students=all_students_stats,
            stats=None,
            lessons=[],
        )

    # פרופיל סטודנט בודד
    if "_students_list" in df.columns:
        mask = df["_students_list"].apply(lambda lst: selected in lst)
        student_df = df[mask].copy()
    else:
        student_df = pd.DataFrame()

    stats = {
        "lessons": len(student_df),
        "hours": round(float(student_df["משך (שעות)"].sum()), 1)
        if "משך (שעות)" in student_df.columns and not student_df.empty
        else 0.0,
        "tutors_count": student_df[COL_TUTOR].nunique() if COL_TUTOR in student_df.columns and not student_df.empty else 0,
    }

    return render_template(
        "students.html",
        students=students,
        selected=selected,
        all_students=[],
        stats=stats,
        lessons=_df_to_records(student_df),
    )


@app.route("/probation")
def probation_page():
    df = _get_df()
    probation_list = cache.get("probation_list")
    if probation_list is None:
        probation_list = load_probation_students()
        cache.set("probation_list", probation_list)

    # Load tutors registry once and build student->assignments map.
    tutors_registry = cache.get("tutors_registry")
    if tutors_registry is None:
        tutors_registry = load_tutors_registry()
        cache.set("tutors_registry", tutors_registry)

    student_assignments: dict[str, list[dict[str, Any]]] = {}
    for tutor in tutors_registry:
        ss = tutor.get("student_subjects", {}) or {}
        default_subj = tutor.get("actual_subject", "") or ", ".join(tutor.get("subjects", []))
        for sname in tutor.get("probation_students", []):
            subj = ss.get(sname) or default_subj
            student_assignments.setdefault(sname, []).append({
                "tutor": tutor["name"],
                "subjects": [subj] if subj else [],
                "phone": tutor.get("phone", ""),
            })

    now = datetime.now()
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    enriched = []
    at_risk_count = 0
    for student in probation_list:
        name = student["name"]
        # שיעורים בכלל
        if "_students_list" in df.columns:
            mask = df["_students_list"].apply(lambda lst, n=name: n in lst)
            student_df = df[mask]
        else:
            student_df = pd.DataFrame()

        total_lessons = len(student_df)
        total_hours = round(float(student_df["משך (שעות)"].sum()), 1) if "משך (שעות)" in student_df.columns and not student_df.empty else 0.0

        # רשימת שיעורים מפורטת
        lessons = []
        if not student_df.empty:
            for _, row in student_df.sort_values(COL_DATE, ascending=False).iterrows():
                lesson = {
                    "date": row[COL_DATE].strftime("%d/%m/%Y") if pd.notna(row.get(COL_DATE)) else "",
                    "tutor": str(row.get(COL_TUTOR, "")),
                    "topic": str(row.get(COL_TOPIC, "")),
                    "start": str(row.get(COL_START, "")),
                    "end": str(row.get(COL_END, "")),
                    "rating": row.get(COL_RATING, ""),
                    "notes": str(row.get(COL_NOTES, "")),
                }
                lessons.append(lesson)

        # שיעורים החודש
        if not student_df.empty and COL_DATE in student_df.columns:
            month_mask = student_df[COL_DATE] >= current_month_start
            lessons_this_month = int(month_mask.sum())
        else:
            lessons_this_month = 0

        # ימים מאז שיעור אחרון
        if not student_df.empty and COL_DATE in student_df.columns and student_df[COL_DATE].notna().any():
            last_lesson_date = student_df[COL_DATE].max()
            days_since = (now - last_lesson_date).days
        else:
            days_since = None

        min_lessons = student["min_lessons"]
        is_at_risk = lessons_this_month < min_lessons if min_lessons > 0 else False
        if is_at_risk:
            at_risk_count += 1

        enriched.append({
            **student,
            "total_lessons": total_lessons,
            "total_hours": total_hours,
            "lessons_this_month": lessons_this_month,
            "days_since": days_since,
            "is_at_risk": is_at_risk,
            "lessons": lessons,
            "assignments": student_assignments.get(name, []),
        })

    # מיון: בסיכון קודם, אח"כ לפי שיעורים החודש
    enriched.sort(key=lambda s: (not s["is_at_risk"], s["lessons_this_month"]))

    all_known = get_all_students(df)

    return render_template(
        "probation.html",
        students=enriched,
        total_probation=len(enriched),
        at_risk_count=at_risk_count,
        current_month=now.strftime("%m/%Y"),
        all_known_students=all_known,
        institutions=config.INSTITUTIONS,
    )


@app.route("/api/probation/add", methods=["POST"])
def api_probation_add():
    data = request.get_json(force=True)
    name = data.get("שם סטודנט", "").strip()
    if not name:
        return jsonify({"error": "שם סטודנט חובה"}), 400
    row = {
        "שם סטודנט": name,
        "סיבה": data.get("סיבה", "").strip(),
        "מוסד": data.get("מוסד", "").strip(),
        "מגמה": data.get("מגמה", "").strip(),
        "תאריך התחלה": data.get("תאריך התחלה", datetime.now().strftime("%d/%m/%Y")),
        "מינימום שיעורים בחודש": str(data.get("מינימום שיעורים בחודש", "4")),
        "הערות": data.get("הערות", "").strip(),
    }
    try:
        append_probation_student(row)
        cache.delete("probation_list")
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/probation/remove", methods=["POST"])
def api_probation_remove():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "שם סטודנט חובה"}), 400
    try:
        removed = remove_probation_student(name)
        cache.delete("probation_list")
        return jsonify({"status": "ok", "removed": removed})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/probation/export")
def probation_export():
    """ייצוא רשימת סטודנטים על תנאי עם השיבוצים שלהם לקובץ Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    probation_list = load_probation_students()
    tutors_registry = load_tutors_registry()
    student_assignments: dict[str, list[dict[str, Any]]] = {}
    for tutor in tutors_registry:
        ss = tutor.get("student_subjects", {}) or {}
        default_subj = tutor.get("actual_subject", "") or ", ".join(tutor.get("subjects", []))
        for sname in tutor.get("probation_students", []):
            subj = ss.get(sname) or default_subj
            student_assignments.setdefault(sname, []).append({
                "tutor": tutor["name"],
                "subjects": subj,
                "phone": tutor.get("phone", ""),
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "סטודנטים על תנאי"
    ws.sheet_view.rightToLeft = True

    headers = ["שם סטודנט", "מוסד", "מגמה", "מקצועות שצריך", "מתגברים משובצים",
               "מקצועות שמכוסים", "מינ' שיעורים בחודש", "תאריך התחלה", "הערות"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6366F1")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for student in probation_list:
        name = student["name"]
        assignments = student_assignments.get(name, [])
        tutors_text = "\n".join(f"{a['tutor']} — {a['subjects']}" for a in assignments) if assignments else ""
        covered = ", ".join(sorted({s.strip() for a in assignments for s in a["subjects"].split(",") if s.strip()}))
        ws.append([
            name,
            student.get("institution", ""),
            student.get("track", ""),
            student.get("reason", ""),
            tutors_text,
            covered,
            student.get("min_lessons", ""),
            student.get("start_date", ""),
            student.get("notes", ""),
        ])

    # Column widths
    widths = [20, 18, 18, 30, 40, 30, 14, 14, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"probation_assignments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/registrations")
def registrations_page():
    """עמוד הרשמות — כל הסטודנטים שנרשמו דרך הטפסים."""
    regs = cache.get("registrations")
    if regs is None:
        regs = load_registrations()
        cache.set("registrations", regs)

    probation_list = cache.get("probation_list")
    if probation_list is None:
        probation_list = load_probation_students()
        cache.set("probation_list", probation_list)
    prob_names = {s["name"] for s in probation_list}

    registry = _get_registry()
    # Build student → tutor assignments for all students (prob + regular)
    student_tutors: dict[str, list[str]] = {}
    for tutor in registry:
        for sname in tutor.get("probation_students", []):
            student_tutors.setdefault(sname, []).append(tutor["name"])
        for sname in tutor.get("regular_students", []):
            student_tutors.setdefault(sname, []).append(tutor["name"])

    enriched = []
    for reg in regs:
        is_prob = reg["name"] in prob_names
        tutors_assigned = student_tutors.get(reg["name"], [])
        enriched.append({
            **reg,
            "is_probation": is_prob,
            "tutors_assigned": tutors_assigned,
        })

    # Stats
    total = len(enriched)
    prob_count = sum(1 for r in enriched if r["is_probation"])
    regular_count = total - prob_count
    assigned_count = sum(1 for r in enriched if r["tutors_assigned"])

    # Group by institution
    by_institution: dict[str, list] = {}
    for r in enriched:
        key = f"{r['institution']} — {r['track']}"
        by_institution.setdefault(key, []).append(r)

    return render_template(
        "registrations.html",
        registrations=enriched,
        by_institution=by_institution,
        total=total,
        prob_count=prob_count,
        regular_count=regular_count,
        assigned_count=assigned_count,
        institutions=config.INSTITUTIONS,
    )


@app.route("/registrations/export")
def registrations_export():
    """ייצוא כל ההרשמות לאקסל."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    regs = load_registrations()
    prob_list = load_probation_students()
    prob_names = {s["name"] for s in prob_list}

    registry = load_tutors_registry()
    student_tutors: dict[str, list[str]] = {}
    for tutor in registry:
        for sname in tutor.get("probation_students", []):
            student_tutors.setdefault(sname, []).append(tutor["name"])
        for sname in tutor.get("regular_students", []):
            student_tutors.setdefault(sname, []).append(tutor["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "הרשמות לתגבורים"
    ws.sheet_view.rightToLeft = True

    headers = ["שם סטודנט", "ת.ז.", "מוסד", "מגמה", "קורסים מבוקשים",
               "סטטוס", "מתגברים משובצים"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6366F1")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for reg in regs:
        is_prob = reg["name"] in prob_names
        tutors = student_tutors.get(reg["name"], [])
        ws.append([
            reg["name"],
            reg["id_num"],
            reg["institution"],
            reg["track"],
            ", ".join(reg["courses"]),
            "על תנאי" if is_prob else "רגיל",
            ", ".join(tutors),
        ])

    widths = [22, 14, 20, 18, 40, 12, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/api/sync-registrations", methods=["POST"])
def api_sync_registrations():
    """סנכרון הרשמות: שיבוץ סטודנטים חדשים למתגברים + עדכון לוז."""
    import auto_scheduler
    import time as _time

    regs = load_registrations()
    registry = load_tutors_registry()
    weekly = load_weekly_schedule()
    prob_list = load_probation_students()
    prob_names = {s["name"] for s in prob_list}

    # Build set of already-assigned students (both prob + regular)
    assigned_set: set[str] = set()
    for t in registry:
        for s in t.get("probation_students", []):
            assigned_set.add(s)
        for s in t.get("regular_students", []):
            assigned_set.add(s)

    # Find unassigned registrations (not probation — those are handled separately)
    new_students = [r for r in regs if r["name"] not in assigned_set and r["name"] not in prob_names]
    if not new_students:
        return jsonify({"status": "ok", "message": "אין סטודנטים חדשים לשיבוץ", "assigned": [], "scheduled": []})

    # Index tutors by (institution, track, subject)
    tutor_by_key: dict[tuple[str, str, str], list[dict]] = {}
    for t in registry:
        inst = t.get("institution", "")
        track = t.get("track", "")
        actual = t.get("actual_subject", "")
        if actual:
            for subj in [s.strip() for s in actual.split(",") if s.strip()]:
                tutor_by_key.setdefault((inst, track, subj), []).append(t)
        for subj in t.get("subjects", []):
            tutor_by_key.setdefault((inst, track, subj), []).append(t)

    # Assign each new student to the best matching tutor
    assigned_log: list[dict] = []
    tutors_changed: set[str] = set()
    for student in new_students:
        matched = False
        for course in student["courses"]:
            key = (student["institution"], student["track"], course)
            candidates = tutor_by_key.get(key, [])
            if not candidates:
                continue
            # Pick tutor with fewest total students
            best = min(candidates, key=lambda t: len(t.get("probation_students", [])) + len(t.get("regular_students", [])))
            # Add student to tutor's regular_students in-memory
            if student["name"] not in best.get("regular_students", []):
                best.setdefault("regular_students", []).append(student["name"])
                best.setdefault("regular_subjects", {})[student["name"]] = course
                tutors_changed.add(best["name"])
                assigned_log.append({
                    "student": student["name"],
                    "tutor": best["name"],
                    "subject": course,
                })
            matched = True
        if not matched:
            assigned_log.append({
                "student": student["name"],
                "tutor": None,
                "subject": None,
                "error": "לא נמצא מתגבר מתאים",
            })

    # Write changes to Google Sheets
    write_count = 0
    for t in registry:
        if t["name"] not in tutors_changed:
            continue
        update_tutor(t["name"], t)
        write_count += 1
        if write_count % 8 == 0:
            _time.sleep(2)

    # Auto-schedule any tutor that now has students but no schedule slot
    scheduled_tutors = {s["tutor"] for s in weekly}
    tutors_needing_schedule = [
        t for t in registry
        if t["name"] in tutors_changed
        and t["name"] not in scheduled_tutors
    ]
    scheduled_log: list[dict] = []
    if tutors_needing_schedule:
        tutor_to_students = {
            t["name"]: list(t.get("probation_students", [])) + list(t.get("regular_students", []))
            for t in registry
        }
        class_schedules = auto_scheduler.load_class_schedules()
        working = list(weekly)
        for t in tutors_needing_schedule:
            slots = auto_scheduler.suggest_for_tutor(t, class_schedules, working, tutor_to_students)
            for slot in slots:
                new_id = append_schedule_slot(slot)
                scheduled_log.append({**slot, "id": new_id})
                working.append(slot)
                _time.sleep(2)

    cache.clear()
    return jsonify({
        "status": "ok",
        "assigned": assigned_log,
        "scheduled": scheduled_log,
        "summary": f"שובצו {len([a for a in assigned_log if a.get('tutor')])} סטודנטים, נוספו {len(scheduled_log)} שיעורים ללוז",
    })


@app.route("/tutors/export")
def tutors_export():
    """ייצוא רשימת מתגברים עם הסטודנטים המשובצים אליהם לקובץ Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tutors_registry = load_tutors_registry()

    wb = Workbook()
    ws = wb.active
    ws.title = "מתגברים ושיבוצים"
    ws.sheet_view.rightToLeft = True

    headers = ["שם מתגבר", "מוסד", "מגמה", "מקצועות (פוטנציאל)", "מקצוע בפועל",
               "סטודנטים על-תנאי משובצים", "סטודנטים רגילים", "סה\"כ סטודנטים", "טלפון", "הערות"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6366F1")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for tutor in tutors_registry:
        students = tutor.get("probation_students", [])
        reg_students = tutor.get("regular_students", [])
        ws.append([
            tutor.get("name", ""),
            tutor.get("institution", ""),
            tutor.get("track", ""),
            ", ".join(tutor.get("subjects", [])),
            tutor.get("actual_subject", ""),
            "\n".join(students),
            "\n".join(reg_students),
            len(students) + len(reg_students),
            tutor.get("phone", ""),
            tutor.get("notes", ""),
        ])

    widths = [22, 18, 18, 30, 22, 35, 35, 12, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"tutors_assignments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/refresh")
def refresh():
    cache.clear()
    return jsonify({"status": "ok"})


# ---------- טופס דיווח ----------

@app.route("/form")
def form_page():
    df = _get_df()
    registry = _get_registry()
    tutors = sorted(t["name"] for t in registry)
    tutor_subjects_map = {t["name"]: t["subjects"] for t in registry}
    students = get_all_students(df)
    return render_template(
        "form.html",
        tutors=tutors,
        students=students,
        tutor_subjects_json=json.dumps(tutor_subjects_map, ensure_ascii=False),
        today=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/api/students")
def api_students():
    df = _get_df()
    return jsonify(get_unique_students(df))


@app.route("/api/tutors")
def api_tutors():
    df = _get_df()
    return jsonify(get_unique_tutors(df))


@app.route("/api/topics")
def api_topics():
    df = _get_df()
    topics = sorted(df["_topic_normalized"].dropna().unique().tolist()) if "_topic_normalized" in df.columns else []
    return jsonify(topics)


# ---------- לוז שבועי ----------

def _week_dates(date_str: str | None = None) -> list[datetime]:
    """מחזיר רשימת 6 תאריכים (א-ו) לשבוע שמכיל את התאריך הנתון."""
    if date_str:
        try:
            base = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            base = datetime.now()
    else:
        base = datetime.now()
    # Python: Monday=0. We want Sunday=0 (יום ראשון)
    # isoweekday: Monday=1..Sunday=7
    iso = base.isoweekday()
    sunday = base - timedelta(days=iso % 7)  # Sunday
    return [sunday + timedelta(days=i) for i in range(6)]  # Sun-Fri


def _build_schedule_matrix(
    weekly: list[dict], onetime: list[dict], reports_df: pd.DataFrame,
    week_dates: list[datetime], tutors_registry: list[dict],
    tutor_filter: str | None = None, institution_filter: str | None = None,
    track_filter: str | None = None,
) -> dict:
    """בונה מטריצה של שיעורים לשבוע עם התאמה לדיווחים."""
    day_names = config.WEEKDAYS  # ["ראשון", "שני", ...]

    # Build tutor info lookup
    tutor_info = {t["name"]: t for t in tutors_registry}

    # Filter by institution/tutor/track if needed
    filtered_tutors = None
    if institution_filter or track_filter:
        filtered_tutors = {t["name"] for t in tutors_registry
                          if (not institution_filter or t["institution"] == institution_filter)
                          and (not track_filter or t["track"] == track_filter)}
    if tutor_filter:
        filtered_tutors = {tutor_filter}

    events: list[dict] = []

    # 1. Weekly schedule slots → expand to this week's dates
    for slot in weekly:
        if filtered_tutors and slot["tutor"] not in filtered_tutors:
            continue
        try:
            day_idx = day_names.index(slot["day"])
        except ValueError:
            continue
        slot_date = week_dates[day_idx]
        info = tutor_info.get(slot["tutor"], {})
        event = {
            "id": slot["id"],
            "type": "weekly",
            "tutor": slot["tutor"],
            "day": slot["day"],
            "date": slot_date.strftime("%Y-%m-%d"),
            "date_display": slot_date.strftime("%d/%m/%Y"),
            "start": slot["start"],
            "end": slot["end"],
            "subject": slot["subject"],
            "institution": info.get("institution", ""),
            "track": info.get("track", ""),
            "probation_students": info.get("probation_students", []),
            "notes": slot["notes"],
            "status": "planned",  # default
            "report": None,
        }
        events.append(event)

    # 2. One-time lessons for this week
    week_start = week_dates[0].date()
    week_end = week_dates[-1].date()
    for lesson in onetime:
        if filtered_tutors and lesson["tutor"] not in filtered_tutors:
            continue
        # Parse date (DD/MM/YYYY or YYYY-MM-DD)
        raw = lesson["date"]
        lesson_date = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                lesson_date = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if not lesson_date or lesson_date < week_start or lesson_date > week_end:
            continue
        day_idx = (lesson_date - week_start).days
        if day_idx < 0 or day_idx >= len(day_names):
            continue
        info = tutor_info.get(lesson["tutor"], {})
        event = {
            "id": lesson["id"],
            "type": "onetime",
            "tutor": lesson["tutor"],
            "day": day_names[day_idx],
            "date": lesson_date.strftime("%Y-%m-%d"),
            "date_display": lesson_date.strftime("%d/%m/%Y"),
            "start": lesson["start"],
            "end": lesson["end"],
            "subject": lesson["subject"],
            "institution": info.get("institution", ""),
            "track": info.get("track", ""),
            "probation_students": info.get("probation_students", []),
            "notes": lesson["notes"],
            "status": "planned",
            "report": None,
        }
        events.append(event)

    # 3. Match with form reports
    if not reports_df.empty and COL_DATE in reports_df.columns and COL_TUTOR in reports_df.columns:
        for event in events:
            event_date = datetime.strptime(event["date"], "%Y-%m-%d").date()
            mask = (
                (reports_df[COL_DATE].dt.date == event_date) &
                (reports_df[COL_TUTOR] == event["tutor"])
            )
            matched = reports_df[mask]
            if not matched.empty:
                report_row = matched.iloc[0]
                actual_start = str(report_row.get(COL_START, ""))
                actual_end = str(report_row.get(COL_END, ""))
                students_str = str(report_row.get(COL_STUDENTS, ""))
                students_list = [s.strip() for s in students_str.replace("،", ",").replace(";", ",").split(",") if s.strip()]
                # Check probation attendance
                prob_attendance = {}
                for ps in event["probation_students"]:
                    prob_attendance[ps] = ps in students_list
                event["status"] = "completed"
                event["report"] = {
                    "start": actual_start,
                    "end": actual_end,
                    "students": students_list,
                    "topic": str(report_row.get(COL_TOPIC, "")),
                    "rating": report_row.get(COL_RATING, ""),
                    "notes": str(report_row.get(COL_NOTES, "")),
                    "probation_attendance": prob_attendance,
                }

    # 4. Find extra reports (not matched to any scheduled lesson)
    matched_keys = set()
    for ev in events:
        if ev["status"] == "completed":
            matched_keys.add((ev["tutor"], ev["date"]))

    if not reports_df.empty and COL_DATE in reports_df.columns:
        for _, report_row in reports_df.iterrows():
            if pd.isna(report_row.get(COL_DATE)):
                continue
            report_date = report_row[COL_DATE].date()
            if report_date < week_start or report_date > week_end:
                continue
            tutor_name = str(report_row.get(COL_TUTOR, ""))
            if filtered_tutors and tutor_name not in filtered_tutors:
                continue
            key = (tutor_name, report_date.strftime("%Y-%m-%d"))
            if key in matched_keys:
                continue
            matched_keys.add(key)
            day_idx = (report_date - week_start).days
            if day_idx < 0 or day_idx >= len(day_names):
                continue
            info = tutor_info.get(tutor_name, {})
            students_str = str(report_row.get(COL_STUDENTS, ""))
            students_list = [s.strip() for s in students_str.replace("،", ",").replace(";", ",").split(",") if s.strip()]
            prob_attendance = {}
            for ps in info.get("probation_students", []):
                prob_attendance[ps] = ps in students_list
            event = {
                "id": None,
                "type": "extra",
                "tutor": tutor_name,
                "day": day_names[day_idx],
                "date": report_date.strftime("%Y-%m-%d"),
                "date_display": report_date.strftime("%d/%m/%Y"),
                "start": str(report_row.get(COL_START, "")),
                "end": str(report_row.get(COL_END, "")),
                "subject": str(report_row.get(COL_TOPIC, "")),
                "institution": info.get("institution", ""),
                "track": info.get("track", ""),
                "probation_students": info.get("probation_students", []),
                "notes": str(report_row.get(COL_NOTES, "")),
                "status": "extra",
                "report": {
                    "start": str(report_row.get(COL_START, "")),
                    "end": str(report_row.get(COL_END, "")),
                    "students": students_list,
                    "topic": str(report_row.get(COL_TOPIC, "")),
                    "rating": report_row.get(COL_RATING, ""),
                    "notes": str(report_row.get(COL_NOTES, "")),
                    "probation_attendance": prob_attendance,
                },
            }
            events.append(event)

    return {
        "events": events,
        "week_dates": [{"name": day_names[i], "date": d.strftime("%Y-%m-%d"), "display": d.strftime("%d/%m")} for i, d in enumerate(week_dates)],
        "hours": list(range(config.SCHEDULE_HOUR_START, config.SCHEDULE_HOUR_END)),
    }


@app.route("/schedule")
def schedule_page():
    date_str = request.args.get("date")
    tutor_filter = request.args.get("tutor")
    institution_filter = request.args.get("institution")
    track_filter = request.args.get("track")
    year_filter = request.args.get("year")

    week_dates = _week_dates(date_str)
    df = _get_df()
    weekly = _get_weekly()
    onetime = _get_onetime()
    registry = _get_registry()
    institutions = config.INSTITUTIONS

    matrix = _build_schedule_matrix(
        weekly, onetime, df, week_dates, registry,
        tutor_filter=tutor_filter, institution_filter=institution_filter,
        track_filter=track_filter,
    )

    academic_events = _get_academic_events(institution_filter, track_filter, year_filter)
    available_years = _available_years(institution_filter, track_filter)

    # Week navigation
    current_sunday = week_dates[0]
    prev_week = (current_sunday - timedelta(days=7)).strftime("%Y-%m-%d")
    next_week = (current_sunday + timedelta(days=7)).strftime("%Y-%m-%d")
    today_str = datetime.now().strftime("%Y-%m-%d")

    tutor_names = sorted(set(t["name"] for t in registry))
    # Build tutor→subjects map for JS
    tutor_subjects_map = {t["name"]: t["subjects"] for t in registry}

    return render_template(
        "schedule.html",
        matrix=matrix,
        prev_week=prev_week,
        next_week=next_week,
        today=today_str,
        institutions=institutions,
        tutor_names=tutor_names,
        tutor_filter=tutor_filter or "",
        institution_filter=institution_filter or "",
        track_filter=track_filter or "",
        year_filter=year_filter or "",
        available_years=available_years,
        weekdays=config.WEEKDAYS,
        schedule_events_json=json.dumps(matrix["events"], ensure_ascii=False, default=str),
        academic_events_json=json.dumps(academic_events, ensure_ascii=False),
        tutor_subjects_json=json.dumps(tutor_subjects_map, ensure_ascii=False),
    )


@app.route("/api/schedule/week")
def api_schedule_week():
    date_str = request.args.get("date")
    tutor_filter = request.args.get("tutor")
    institution_filter = request.args.get("institution")
    track_filter = request.args.get("track")
    year_filter = request.args.get("year")

    week_dates = _week_dates(date_str)
    df = _get_df()
    weekly = _get_weekly()
    onetime = _get_onetime()
    registry = _get_registry()

    matrix = _build_schedule_matrix(
        weekly, onetime, df, week_dates, registry,
        tutor_filter=tutor_filter, institution_filter=institution_filter,
        track_filter=track_filter,
    )
    matrix["academic_events"] = _get_academic_events(institution_filter, track_filter, year_filter)
    return jsonify(matrix)


@app.route("/class-schedule/print")
def class_schedule_print():
    """Printable / PDF-export view of the academic class schedule for a
    specific institution + track (+ optional year)."""
    institution = request.args.get("institution", "").strip()
    track = request.args.get("track", "").strip()
    year = request.args.get("year", "").strip() or None
    if not institution or not track:
        return redirect("/schedule")

    events = _get_academic_events(institution, track, year)
    available_years = _available_years(institution, track)

    # Trim hour range to actual event bounds (with 1-hour padding).
    if events:
        starts = [int(ev["start"].split(":")[0]) for ev in events if ev.get("start")]
        ends = [int(ev["end"].split(":")[0]) + (1 if ev["end"].split(":")[1] != "00" else 0)
                for ev in events if ev.get("end")]
        hour_start = max(0, min(starts))
        hour_end = min(24, max(ends) + 1)
    else:
        hour_start = config.SCHEDULE_HOUR_START
        hour_end = config.SCHEDULE_HOUR_END

    return render_template(
        "class_schedule_print.html",
        institution=institution,
        track=track,
        year=year or "",
        available_years=available_years,
        events=events,
        weekdays=config.WEEKDAYS,
        hours=list(range(hour_start, hour_end)),
    )


@app.route("/api/schedule/add", methods=["POST"])
def api_schedule_add():
    data = request.get_json(force=True)
    lesson_type = data.get("type", "weekly")
    tutor = data.get("tutor", "").strip()
    if not tutor:
        return jsonify({"error": "שם מתגבר חובה"}), 400

    # Validate tutor exists in registry
    registry = _get_registry()
    tutor_record = next((t for t in registry if t["name"] == tutor), None)
    if not tutor_record:
        return jsonify({"error": f"המתגבר '{tutor}' לא רשום במערכת"}), 400

    # Validate subject is in tutor's subjects
    subject = data.get("subject", "").strip()
    if subject and tutor_record["subjects"] and subject not in tutor_record["subjects"]:
        return jsonify({"error": f"המקצוע '{subject}' לא משויך למתגבר '{tutor}'"}), 400

    try:
        if lesson_type == "onetime":
            new_id = append_onetime_lesson({
                "tutor": tutor,
                "date": data.get("date", ""),
                "start": data.get("start", ""),
                "end": data.get("end", ""),
                "subject": data.get("subject", ""),
                "notes": data.get("notes", ""),
            })
        else:
            new_id = append_schedule_slot({
                "tutor": tutor,
                "day": data.get("day", ""),
                "start": data.get("start", ""),
                "end": data.get("end", ""),
                "subject": data.get("subject", ""),
                "notes": data.get("notes", ""),
            })
        cache.delete("weekly")
        cache.delete("onetime")
        return jsonify({"status": "ok", "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/schedule/update", methods=["POST"])
def api_schedule_update():
    data = request.get_json(force=True)
    lesson_type = data.get("type", "weekly")
    lesson_id = str(data.get("id", "")).strip()
    if not lesson_id:
        return jsonify({"error": "ID חובה"}), 400

    try:
        if lesson_type == "onetime":
            ok = update_onetime_lesson(lesson_id, {
                "tutor": data.get("tutor", ""),
                "date": data.get("date", ""),
                "start": data.get("start", ""),
                "end": data.get("end", ""),
                "subject": data.get("subject", ""),
                "notes": data.get("notes", ""),
            })
        else:
            ok = update_schedule_slot(lesson_id, {
                "tutor": data.get("tutor", ""),
                "day": data.get("day", ""),
                "start": data.get("start", ""),
                "end": data.get("end", ""),
                "subject": data.get("subject", ""),
                "notes": data.get("notes", ""),
            })
        cache.delete("weekly")
        cache.delete("onetime")
        return jsonify({"status": "ok", "updated": ok})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/schedule/remove", methods=["POST"])
def api_schedule_remove():
    data = request.get_json(force=True)
    lesson_type = data.get("type", "weekly")
    lesson_id = str(data.get("id", "")).strip()
    if not lesson_id:
        return jsonify({"error": "ID חובה"}), 400

    try:
        if lesson_type == "onetime":
            ok = remove_onetime_lesson(lesson_id)
        else:
            ok = remove_schedule_slot(lesson_id)
        cache.delete("weekly")
        cache.delete("onetime")
        if not ok:
            return jsonify({"error": f"שיעור עם ID {lesson_id} לא נמצא"}), 404
        return jsonify({"status": "ok", "removed": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------- ניהול מתגברים ----------

@app.route("/tutors-registry")
def tutors_registry_page():
    registry = _get_registry()
    institutions = config.INSTITUTIONS
    probation_list = cache.get("probation_list")
    if probation_list is None:
        probation_list = load_probation_students()
        cache.set("probation_list", probation_list)
    # Group weekly slots per tutor for the "שיעורים בשבוע" column.
    weekly = _get_weekly()
    weekday_order = {d: i for i, d in enumerate(config.WEEKDAYS)}
    tutor_lessons: dict[str, list[dict]] = {}
    for slot in weekly:
        tutor_lessons.setdefault(slot["tutor"], []).append(slot)
    for slots in tutor_lessons.values():
        slots.sort(key=lambda s: (weekday_order.get(s.get("day", ""), 99), s.get("start", "")))
    return render_template(
        "tutors_registry.html",
        tutors=registry,
        institutions=institutions,
        probation_students=probation_list,
        tutor_lessons=tutor_lessons,
    )


@app.route("/api/auto-schedule", methods=["POST"])
def api_auto_schedule():
    """שיבוץ אוטומטי. mode=preview מחזיר הצעות; mode=commit שומר לסליל."""
    import auto_scheduler
    data = request.get_json(silent=True) or {}
    mode = data.get("mode", "preview")
    registry = _get_registry()
    weekly = _get_weekly()
    suggestions = auto_scheduler.suggest_all(registry, weekly, only_unscheduled=True)
    if mode != "commit":
        return jsonify({"suggestions": suggestions, "count": len(suggestions)})
    saved = []
    try:
        for s in suggestions:
            new_id = append_schedule_slot({
                "tutor": s["tutor"],
                "day": s["day"],
                "start": s["start"],
                "end": s["end"],
                "subject": s["subject"],
                "notes": s.get("notes", ""),
            })
            saved.append({**s, "id": new_id})
        cache.delete("weekly")
        return jsonify({"status": "ok", "saved": saved, "count": len(saved)})
    except Exception as e:
        return jsonify({"error": str(e), "saved": saved}), 500


@app.route("/schedule/export")
def schedule_export():
    """ייצוא Excel של כל השיעורים השבועיים בכל המוסדות עם המתגבר, היום, השעות והמקצוע."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    weekly = load_weekly_schedule()
    registry = load_tutors_registry()
    tutor_info = {t["name"]: t for t in registry}
    weekday_order = {d: i for i, d in enumerate(config.WEEKDAYS)}

    rows: list[dict] = []
    for slot in weekly:
        info = tutor_info.get(slot["tutor"], {})
        prob_students = info.get("probation_students", [])
        reg_students = info.get("regular_students", [])
        all_names = list(prob_students) + list(reg_students)
        rows.append({
            "institution": info.get("institution", ""),
            "track": info.get("track", ""),
            "tutor": slot["tutor"],
            "day": slot.get("day", ""),
            "start": slot.get("start", ""),
            "end": slot.get("end", ""),
            "subject": slot.get("subject", "") or info.get("actual_subject", ""),
            "students": "\n".join(all_names),
            "num_students": len(prob_students) + len(reg_students),
            "notes": slot.get("notes", ""),
        })
    rows.sort(key=lambda r: (r["institution"], r["track"], r["tutor"],
                             weekday_order.get(r["day"], 99), r["start"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "שיעורים שבועיים"
    ws.sheet_view.rightToLeft = True

    headers = ["מוסד", "מגמה", "שם מתגבר", "יום", "התחלה", "סיום",
               "מקצוע", "סטודנטים", "מס' סטודנטים"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6366F1")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r["institution"], r["track"], r["tutor"], r["day"],
            r["start"], r["end"], r["subject"],
            r["students"], r["num_students"],
        ])

    widths = [20, 18, 22, 10, 10, 10, 22, 40, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"weekly_schedule_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/api/tutors/add", methods=["POST"])
def api_tutors_add():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "שם מתגבר חובה"}), 400
    try:
        append_tutor({
            "name": name,
            "institution": data.get("institution", ""),
            "track": data.get("track", ""),
            "subjects": data.get("subjects", []),
            "probation_students": data.get("probation_students", []),
            "phone": data.get("phone", ""),
            "notes": data.get("notes", ""),
        })
        cache.delete("registry")
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tutors/update", methods=["POST"])
def api_tutors_update():
    data = request.get_json(force=True)
    original_name = data.get("original_name", "").strip()
    if not original_name:
        return jsonify({"error": "שם מתגבר מקורי חובה"}), 400
    try:
        ok = update_tutor(original_name, {
            "name": data.get("name", original_name),
            "institution": data.get("institution", ""),
            "track": data.get("track", ""),
            "subjects": data.get("subjects", []),
            "probation_students": data.get("probation_students", []),
            "phone": data.get("phone", ""),
            "notes": data.get("notes", ""),
        })
        cache.delete("registry")
        return jsonify({"status": "ok", "updated": ok})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tutors/remove", methods=["POST"])
def api_tutors_remove():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "שם מתגבר חובה"}), 400
    try:
        ok = remove_tutor(name)
        remove_schedule_by_tutor(name)
        cache.delete("registry")
        cache.delete("weekly")
        cache.delete("onetime")
        return jsonify({"status": "ok", "removed": ok})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tutors/subjects")
def api_tutors_subjects():
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify([])
    return jsonify(get_tutor_subjects(name))


@app.route("/api/submit", methods=["POST"])
def api_submit():
    data = request.get_json(force=True)
    required = [COL_TUTOR, COL_DATE, COL_START, COL_END, COL_STUDENTS, COL_TOPIC]
    for field in required:
        if not data.get(field):
            return jsonify({"error": f"שדה חובה חסר: {field}"}), 400

    from datetime import datetime as dt

    # Convert date from ISO (2026-05-19) to DD/MM/YYYY format
    raw_date = data[COL_DATE].strip()
    try:
        parsed_date = dt.strptime(raw_date, "%Y-%m-%d")
        formatted_date = parsed_date.strftime("%d/%m/%Y")
    except ValueError:
        formatted_date = raw_date

    # Ensure time has seconds (HH:MM -> HH:MM:SS)
    raw_start = data[COL_START].strip()
    raw_end = data[COL_END].strip()
    formatted_start = raw_start + ":00" if raw_start.count(":") == 1 else raw_start
    formatted_end = raw_end + ":00" if raw_end.count(":") == 1 else raw_end

    row = {
        COL_TUTOR: data[COL_TUTOR].strip(),
        COL_DATE: formatted_date,
        COL_START: formatted_start,
        COL_END: formatted_end,
        COL_STUDENTS: data[COL_STUDENTS].strip(),
        COL_TOPIC: data[COL_TOPIC].strip(),
        COL_RATING: str(data.get(COL_RATING, "")),
        COL_NOTES: data.get(COL_NOTES, "").strip(),
        COL_TIMESTAMP: dt.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    try:
        append_row(row)
        cache.delete("df")
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import os
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(debug=debug, port=int(os.environ.get("PORT", 5003)))
